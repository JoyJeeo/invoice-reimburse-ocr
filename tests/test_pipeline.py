from pathlib import Path

from openpyxl import load_workbook

from invoice_reimburse_ocr.models import REIMBURSEMENT_COLUMNS
from invoice_reimburse_ocr.ocr import SidecarTextOCREngine
from invoice_reimburse_ocr.pipeline import process_folder


def test_process_folder_exports_clean_reimbursement_file(tmp_path: Path):
    input_dir = tmp_path / "inputs"
    input_dir.mkdir()
    image_file = input_dir / "invoice.jpg"
    image_file.write_bytes(b"fake image bytes")
    image_file.with_suffix(".txt").write_text(
        """
        发票代码: 031002300111
        发票号码: 12345678
        开票日期: 2026-06-17
        合计金额: 100.50
        合计税额: 6.03
        价税合计: 106.53
        购买方名称: 上海测试科技有限公司
        购买方纳税人识别号: 91310000MA1K000000
        销售方名称: 北京服务有限公司
        销售方纳税人识别号: 91110000MA1K111111
        """,
        encoding="utf-8",
    )

    result = process_folder(input_dir, tmp_path, SidecarTextOCREngine())

    assert result.total == 1
    assert result.review_count == 0
    assert result.template_file.exists()
    assert result.reimbursement_file.exists()
    assert result.log_file.exists()

    wb = load_workbook(result.reimbursement_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == REIMBURSEMENT_COLUMNS
    assert "原始文件名" not in headers
    assert "识别状态" not in headers
    assert ws.cell(row=2, column=1).value == "发票"
    assert ws.cell(row=2, column=9).value == 106.53


def test_process_folder_exports_currency_columns(tmp_path: Path):
    input_dir = tmp_path / "inputs"
    input_dir.mkdir()
    image_file = input_dir / "invoice.jpg"
    image_file.write_bytes(b"fake image bytes")
    image_file.with_suffix(".txt").write_text(
        """
        Currency: USD
        发票代码: 031002300111
        发票号码: 12345678
        开票日期: 2026-06-17
        合计金额: USD 100.00
        合计税额: USD 6.00
        价税合计: USD 106.00
        购买方名称: 上海测试科技有限公司
        购买方纳税人识别号: 91310000MA1K000000
        销售方名称: 北京服务有限公司
        销售方纳税人识别号: 91110000MA1K111111
        """,
        encoding="utf-8",
    )

    result = process_folder(input_dir, tmp_path, SidecarTextOCREngine(), lambda currency: 7.2)

    wb = load_workbook(result.reimbursement_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers[9:13] == ["币种", "原币种费用", "汇率", "人民币费用"]
    assert ws.cell(row=2, column=10).value == "USD"
    assert ws.cell(row=2, column=11).value == 106.0
    assert ws.cell(row=2, column=12).value == 7.2
    assert ws.cell(row=2, column=13).value == 763.2


def test_process_folder_exports_payment_records_and_deduplicates(tmp_path: Path):
    input_dir = tmp_path / "inputs"
    input_dir.mkdir()
    first = input_dir / "payment_1.jpg"
    second = input_dir / "payment_2.jpg"
    first.write_bytes(b"fake image bytes")
    second.write_bytes(b"fake image bytes")
    payment_text = """
    支付凭证
    交易单号: PAY202606170001
    付款日期: 2026-06-17
    付款金额: ￥88.66
    付款方: 上海测试科技有限公司
    收款方: 北京服务有限公司
    支付方式: 银行转账
    """
    first.with_suffix(".txt").write_text(payment_text, encoding="utf-8")
    second.with_suffix(".txt").write_text(payment_text, encoding="utf-8")

    result = process_folder(input_dir, tmp_path, SidecarTextOCREngine())

    wb = load_workbook(result.reimbursement_file)
    ws = wb.active
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "付款记录"
    assert ws.cell(row=2, column=5).value == "2026-06-17"
    assert ws.cell(row=2, column=6).value == "PAY202606170001"
    assert ws.cell(row=2, column=9).value == 88.66
    assert ws.cell(row=2, column=18).value == "上海测试科技有限公司"
    assert ws.cell(row=2, column=19).value == "北京服务有限公司"

    log_wb = load_workbook(result.log_file)
    log_ws = log_wb.active
    statuses = [log_ws.cell(row=row, column=3).value for row in range(2, log_ws.max_row + 1)]
    assert "重复已去重" in statuses
