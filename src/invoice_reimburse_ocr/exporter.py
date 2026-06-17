from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import InvoiceRecord, REIMBURSEMENT_COLUMNS


def create_run_output_dir(base_dir: Path, timestamp: datetime | None = None) -> Path:
    timestamp = timestamp or datetime.now()
    output_dir = base_dir / "outputs" / timestamp.strftime("%Y%m%d_%H%M%S")
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def export_template(output_dir: Path) -> Path:
    path = output_dir / "发票报销模板.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "发票报销模板"
    ws.append(REIMBURSEMENT_COLUMNS)
    _format_sheet(ws)
    wb.save(path)
    return path


def export_reimbursement(records: list[InvoiceRecord], output_dir: Path, timestamp: datetime | None = None) -> Path:
    timestamp = timestamp or datetime.now()
    path = output_dir / f"报销明细_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "报销明细"
    ws.append(REIMBURSEMENT_COLUMNS)
    for record in records:
        row = record.to_reimbursement_row()
        ws.append([row[column] for column in REIMBURSEMENT_COLUMNS])
    _format_sheet(ws)
    _format_amount_columns(ws, [4, 5, 6])
    wb.save(path)
    return path


def export_processing_log(records: list[InvoiceRecord], output_dir: Path) -> Path:
    path = output_dir / "处理日志.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "处理日志"
    headers = ["原始文件名", "识别状态", "错误信息", "发票号码"]
    ws.append(headers)
    for record in records:
        ws.append([record.source_file, record.status, "；".join(record.errors), record.invoice_number])
    _format_sheet(ws)
    wb.save(path)
    return path


def _format_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 36)
    ws.freeze_panes = "A2"


def _format_amount_columns(ws, columns: list[int]) -> None:
    for column in columns:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=column).number_format = "0.00"
